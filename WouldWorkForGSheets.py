import openpyxl as openpyxl
from PIL import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import gspread

if __name__ == '__main__':
    imageName = input("Please enter an image name:")
    with Image.open(imageName) as inputImage:
        # inputImage.show()
        rgbInputImage = inputImage.convert('RGB')
        width = inputImage.width
        height = inputImage.height

    sa = gspread.service_account()
    sh = sa.open("Image to Sheet")
    worksheet = sh.add_worksheet(title="Image", rows=height, cols=width)

    wks = sh.worksheet("Image")
    for hPixel in range(1, height):
        shift = hPixel - 1 if hPixel > 1 else hPixel

        vals = [item for item in wks.col_values(1) if item]
        index = len(vals)+1 - shift

        for wPixel in range(1, width):
            r, g, b = rgbInputImage.getpixel((wPixel, hPixel))
            columnLetter = get_column_letter(wPixel)

            rVal = r/255
            gVal = g/255
            bVal = b/255

            wks.format(columnLetter + str(index + hPixel), {
                "backgroundColor": {
                    "red": rVal,
                    "green": 0.0,
                    "blue": 0.0
                }})
            wks.format(columnLetter + str(index + hPixel + 1), {
                "backgroundColor": {
                    "red": 0.0,
                    "green": gVal,
                    "blue": 0.0
                }})
            wks.format(columnLetter + str(index + hPixel + 2), {
                "backgroundColor": {
                    "red": 0.0,
                    "green": 0.0,
                    "blue": bVal
                }})
