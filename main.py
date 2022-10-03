import shutil
import openpyxl as openpyxl
from PIL import Image
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import cv2
import moviepy.editor as mp
import lxml


def splitVideo(video):
    vidcap = cv2.VideoCapture(video)
    success, image = vidcap.read()
    count = 0
    while success:
        cv2.imwrite("Images/%d.jpg" % count, image)
        success, image = vidcap.read()
        count += 1


def deleteFilesInDirectory(directory):
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))


def videoOptions():
    video = input("Please enter video name:\n->  ")
    resizeYN = input("Would you like to resize the video(recommended size is 144p)?:\n(y/n)->  ")
    deleteFilesInDirectory('Images')
    if resizeYN == "y" or resizeYN == "Y":
        # do the resize
        resizeW = ""
        resizeH = ""
        while not resizeW.isdigit():
            resizeW = input("Please enter the width:\n->  ")
        while not resizeH.isdigit():
            resizeH = input("Please enter the height:\n->  ")
        clip = mp.VideoFileClip(video)
        clip_resized = clip.resize(height=int(resizeH))
        clip_resized = clip_resized.resize(width=int(resizeW))
        clip_resized.write_videofile("resized-" + video)
        splitVideo("resized-" + video)
    else:
        splitVideo(video)


if __name__ == '__main__':

    lameMode = True
    wb = openpyxl.Workbook(write_only=lameMode)
    dontCont = True

    while dontCont:
        val = input("1: use new video \n2: use existing photos in Image folder\n->  ")
        if val == '1' or val == '2':
            dontCont = False
        else:
            print('Sorry, please enter a valid input')

    # sample video: 'dolphin21-preview_COkhxs4Y.mp4'

    if val == '1':
        videoOptions()

    images = os.listdir('Images')

    for i in range(len(images)):
        print(i)

        wb.create_sheet(str(i))
        wb.active = wb[str(i)]
        ws = wb.active
        ws.title = str(i)

        with Image.open('Images/' + str(i) + '.jpg') as inputImage:

            rgbInputImage = inputImage.convert('RGB')
            width = inputImage.width
            height = inputImage.height

        if lameMode:
            wPixel = 0
            hPixel = 0

            for hPixel in range(1, height):
                tempList = []
                for wPixel in range(1, width):
                    r, g, b = rgbInputImage.getpixel((wPixel, hPixel))
                    rgbVal = '=@myRGB('+str(r)+','+str(g)+','+str(b)+')'
                    cell = WriteOnlyCell(ws, value=rgbVal)
                    # cell.fill = PatternFill("solid", fgColor=rgbVal)
                    tempList.append(cell)
                    wPixel += 1
                hPixel += 1
                ws.append([c for c in tempList])


        else:
            for hPixel in range(1, height):
                index = (hPixel * 3) - 2
                for wPixel in range(1, width):
                    r, g, b = rgbInputImage.getpixel((wPixel, hPixel))
                    columnLetter = get_column_letter(wPixel)

                    rVal = 'FF%02x%02x%02x' % (r, 0, 0)
                    gVal = 'FF%02x%02x%02x' % (0, g, 0)
                    bVal = 'FF%02x%02x%02x' % (0, 0, b)

                    ws[columnLetter + str(index)].fill = PatternFill(start_color=rVal, end_color=rVal,
                                                                     fill_type="solid")
                    ws[columnLetter + str(index + 1)].fill = PatternFill(start_color=gVal, end_color=gVal,
                                                                         fill_type="solid")
                    ws[columnLetter + str(index + 2)].fill = PatternFill(start_color=bVal, end_color=bVal,
                                                                         fill_type="solid")
        ws.sheet_view.zoomScale = 10
        ws.sheet_view.showGridLines = False
        # TODO:add a reduced ram usage mode

        # wb.save(filename='sample_book.xlsx') in every loop reduces ram used but is slower.
    wb.save(filename='video.xlsx')
    # webbrowser.open('sample_book.xlsx')
